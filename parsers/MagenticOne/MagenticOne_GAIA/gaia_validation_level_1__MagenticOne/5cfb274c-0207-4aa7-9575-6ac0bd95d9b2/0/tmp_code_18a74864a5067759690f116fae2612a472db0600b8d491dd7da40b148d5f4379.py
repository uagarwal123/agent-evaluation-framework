import openpyxl
from openpyxl.styles import PatternFill

def is_green(cell):
    """Check if a cell's fill matches the predefined green color for Earl's plots."""
    return cell.fill.start_color.index == 'FF00FF00'

def list_green_cells(filename):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active
    green_cells = []

    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            if is_green(cell):
                green_cells.append((cell.row, cell.column))

    return green_cells

filename = '5cfb274c-0207-4aa7-9575-6ac0bd95d9b2.xlsx'
green_cells = list_green_cells(filename)
print("Green cells (Earl's plots):", green_cells)
