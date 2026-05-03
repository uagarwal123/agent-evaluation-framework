from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill

# Load the Excel file
workbook = load_workbook(filename='65afbc8a-89ca-4ad5-8d62-355bb401f61d.xlsx')
sheet = workbook.active

# Define the starting position
start_pos = (1, 1)  # Starting in cell A1, which is 1-indexed in openpyxl

# Directions for movement: down, right, up, left
directions = [(2, 0), (0, 2), (-2, 0), (0, -2)]
current_pos = start_pos
moves = 0

# Function to get the fill color of a cell as a hex code
def get_fill_color_as_hex(cell):
    fill = cell.fill
    if hasattr(fill, 'fgColor') and fill.fgColor.type == 'rgb':
        rgb = fill.fgColor.rgb
        # Return None if the RGB value is not properly formed
        if rgb is not None and len(rgb) == 8:  # Expecting a value like 'FF0000FF' including alpha
            return rgb[2:]  # Removing the first two characters ('FF') from ARGB format
    return None

# Function to determine if a cell can be moved to (not blue)
def can_move_to(cell):
    color = get_fill_color_as_hex(cell)
    return color != '0000FF'  # Assuming '0000FF' is the hex code for blue

# Find the 11th move position
try:
    while moves < 11:
        for direction in directions:
            next_row = current_pos[0] + direction[0]
            next_col = current_pos[1] + direction[1]
            if 1 <= next_row <= sheet.max_row and 1 <= next_col <= sheet.max_column:
                next_cell = sheet.cell(row=next_row, column=next_col)
                if can_move_to(next_cell):
                    current_pos = (next_row, next_col)
                    moves += 1
                    if moves == 11:
                        break

except IndexError:
    print("Path could not be completed within bounds and avoiding blue cells.")

# Get the color of the final cell
final_cell = sheet.cell(row=current_pos[0], column=current_pos[1])
final_color_hex = get_fill_color_as_hex(final_cell)

if final_color_hex:
    print(f"The color of the cell after the 11th move is: {final_color_hex} (hex)")
else:
    print("No valid hex color code found for the final cell.")
